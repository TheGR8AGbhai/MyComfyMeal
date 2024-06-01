from flask import Flask, request, render_template
from openpyxl import Workbook
import os

app = Flask(__name__)

# Ensure the Excel file exists or create it
def ensure_excel_exists(filename):
    if not os.path.exists(filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Email", "Phone", "Message"])
        workbook.save(filename)

EXCEL_FILE = 'contacts.xlsx'
ensure_excel_exists(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form.get('name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    message = request.form.get('message')

    if name and email and phone and message:
        workbook = Workbook()
        workbook = Workbook()
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        sheet.append([name, email, phone, message])
        workbook.save(EXCEL_FILE)
        return 'Form submitted successfully!'
    return 'Please fill all fields!', 400

if __name__ == '__main__':
    app.run(debug=True)
